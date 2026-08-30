#!/usr/bin/env python3
"""Probe public publication artifacts without committing third-party source files.

Downloads are ephemeral. The script records stable landing URLs, resolved artifact
URLs, byte sizes, SHA-256 digests, workbook/archive structure, formulas, and small
cell contexts around bounded target values. It never commits or republishes the
third-party source artifacts themselves.
"""

from __future__ import annotations

import hashlib
import io
import json
import re
import zipfile
from dataclasses import asdict, dataclass
from typing import Iterable
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

UA = "metabolic-engineering-forensics/0.1 (+research artifact verification)"
TIMEOUT = 60
MAX_HITS = 100
CONTEXT_RADIUS = 2


@dataclass
class Probe:
    case: str
    label: str
    landing_url: str
    resolved_url: str | None = None
    status_code: int | None = None
    content_type: str | None = None
    size_bytes: int | None = None
    sha256: str | None = None
    artifact_kind: str | None = None
    workbook_sheets: list[dict] | None = None
    archive_members: list[str] | None = None
    target_hits: list[dict] | None = None
    error: str | None = None


def get(url: str) -> requests.Response:
    r = requests.get(url, headers={"User-Agent": UA}, timeout=TIMEOUT, allow_redirects=True)
    r.raise_for_status()
    return r


def find_download(landing_url: str, anchor_pattern: str) -> str:
    r = get(landing_url)
    soup = BeautifulSoup(r.text, "html.parser")
    pattern = re.compile(anchor_pattern, re.I)
    candidates: list[tuple[str, str]] = []
    for a in soup.find_all("a", href=True):
        text = " ".join(a.stripped_strings)
        href = a.get("href")
        if pattern.search(text):
            candidates.append((text, urljoin(r.url, href)))
    if not candidates:
        raise RuntimeError(f"no anchor matching {anchor_pattern!r} at {landing_url}")
    candidates.sort(key=lambda x: ("download" not in x[0].lower(), len(x[0])))
    return candidates[0][1]


def looks_like_xlsx(data: bytes, resolved_url: str, content_type: str | None) -> bool:
    path = urlparse(resolved_url).path.lower()
    if path.endswith(".xlsx"):
        return True
    ct = (content_type or "").lower()
    if "spreadsheetml" in ct or "ms-excel" in ct:
        return True
    if not zipfile.is_zipfile(io.BytesIO(data)):
        return False
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            return "xl/workbook.xml" in zf.namelist()
    except zipfile.BadZipFile:
        return False


def normalized(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return format(value, ".12g")
    return str(value).strip()


def target_match(value: object, targets: Iterable[str]) -> bool:
    sv = normalized(value)
    for raw in targets:
        t = str(raw).strip()
        if not t:
            continue
        if sv == t:
            return True
        # Numeric equality protects against 218 versus 218.0.
        try:
            if abs(float(sv) - float(t)) <= 1e-9:
                return True
        except (TypeError, ValueError):
            pass
        if len(t) >= 4 and t.lower() in sv.lower():
            return True
    return False


def cell_context(ws, row: int, col: int, radius: int = CONTEXT_RADIUS) -> list[dict]:
    out: list[dict] = []
    for r in range(max(1, row - radius), min(ws.max_row, row + radius) + 1):
        vals = []
        for c in range(max(1, col - radius), min(ws.max_column, col + radius) + 1):
            cell = ws.cell(r, c)
            vals.append({"cell": cell.coordinate, "value": normalized(cell.value)[:300]})
        out.append({"row": r, "cells": vals})
    return out


def workbook_inventory(data: bytes, targets: Iterable[str]) -> tuple[list[dict], list[dict]]:
    wb_formula = load_workbook(io.BytesIO(data), read_only=False, data_only=False)
    wb_values = load_workbook(io.BytesIO(data), read_only=False, data_only=True)
    sheets: list[dict] = []
    hits: list[dict] = []
    for ws in wb_formula.worksheets:
        sheets.append({"title": ws.title, "max_row": ws.max_row, "max_column": ws.max_column})
        ws_values = wb_values[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cached = ws_values[cell.coordinate].value
                if target_match(cell.value, targets) or target_match(cached, targets):
                    hit = {
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "value": normalized(cell.value)[:300],
                        "cached_value": normalized(cached)[:300],
                        "formula": normalized(cell.value)[:300] if isinstance(cell.value, str) and cell.value.startswith("=") else None,
                        "context": cell_context(ws_values, cell.row, cell.column),
                    }
                    hits.append(hit)
                    if len(hits) >= MAX_HITS:
                        return sheets, hits
    return sheets, hits


def probe_download(case: str, label: str, landing: str, download_url: str, targets: list[str]) -> Probe:
    p = Probe(case=case, label=label, landing_url=landing)
    try:
        r = get(download_url)
        data = r.content
        p.resolved_url = r.url
        p.status_code = r.status_code
        p.content_type = r.headers.get("content-type")
        p.size_bytes = len(data)
        p.sha256 = hashlib.sha256(data).hexdigest()

        if looks_like_xlsx(data, r.url, p.content_type):
            p.artifact_kind = "xlsx"
            sheets, hits = workbook_inventory(data, targets)
            p.workbook_sheets = sheets
            p.target_hits = hits or None
        elif zipfile.is_zipfile(io.BytesIO(data)):
            p.artifact_kind = "zip"
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                # Only report logical file members, not internal Office ZIP contents.
                p.archive_members = [name for name in zf.namelist() if not name.endswith("/")]
                all_sheets: list[dict] = []
                all_hits: list[dict] = []
                for name in p.archive_members:
                    if name.lower().endswith(".xlsx"):
                        try:
                            sheets, hits = workbook_inventory(zf.read(name), targets)
                            for item in sheets:
                                item["archive_member"] = name
                            for item in hits:
                                item["archive_member"] = name
                            all_sheets.extend(sheets)
                            all_hits.extend(hits)
                        except Exception as exc:
                            all_sheets.append({"archive_member": name, "error": str(exc)})
                p.workbook_sheets = all_sheets or None
                p.target_hits = all_hits or None
        else:
            p.artifact_kind = "other"
    except Exception as exc:
        p.error = f"{type(exc).__name__}: {exc}"
    return p


def main() -> int:
    specs = [
        {
            "case": "kim-2019",
            "label": "Supplementary Dataset 1 — raw fed-batch fermentations",
            "landing": "https://www.nature.com/articles/s41589-019-0295-5",
            "anchor": r"Supplementary Dataset 1",
            "targets": ["50.2", "47.2", "46.7", "49.2", "50.1", "ROP1_34", "FFA"],
        },
        {
            "case": "park-2022",
            "label": "Source Data Fig. 7 — final lutein fed-batch",
            "landing": "https://www.nature.com/articles/s41929-022-00820-4",
            "anchor": r"Source Data Fig\. 7",
            "targets": ["218", "218.0", "5.01", "LUT5MH1", "lutein", "productivity"],
        },
        {
            "case": "cho-2026",
            "label": "Figshare source-data archive",
            "landing": "https://springernature.figshare.com/articles/dataset/Cho_and_Prabowo-etal-Source_data/29264624",
            "anchor": r"Download",
            "targets": ["141.5", "141.517", "2.95", "SC97", "1,3-PDO", "productivity"],
        },
    ]

    results: list[Probe] = []
    for spec in specs:
        try:
            url = find_download(spec["landing"], spec["anchor"])
            result = probe_download(spec["case"], spec["label"], spec["landing"], url, spec["targets"])
        except Exception as exc:
            result = Probe(
                case=spec["case"], label=spec["label"], landing_url=spec["landing"],
                error=f"{type(exc).__name__}: {exc}",
            )
        results.append(result)

    payload = {
        "schema_version": 2,
        "principle": "Downloaded third-party bytes are ephemeral; only provenance metadata, workbook structure, formulas, and bounded cell contexts are emitted.",
        "probes": [asdict(p) for p in results],
    }
    text = json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=True)
    print(text)
    with open("probe-results.json", "w", encoding="utf-8") as fh:
        fh.write(text + "\n")
    return 0 if all(p.error is None for p in results) else 1


if __name__ == "__main__":
    raise SystemExit(main())
